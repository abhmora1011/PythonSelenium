import openpyxl

'''READ INFO TO FILE'''
book = openpyxl.load_workbook("/Users/hnstabe/Downloads/PythonDemo.xlsx")

dict = {}

sheet = book.active

cell = sheet.cell(row=1, column=2)

print(cell.value)

'''TO WRITE TO FILE'''
sheet.cell(row=3, column=2).value = "Clarisse"

print(sheet.cell(row=3, column=2).value)

print(sheet.max_row) # read row

print(sheet.max_column) # read column

#   print(sheet['A1'].value)

# for i in range(1, sheet.max_row + 1):
#     if sheet.cell(row=i, column=1).value == "Testcase1":
#         for j in range(1, sheet.max_column + 1):
#             print(sheet.cell(row=i, column=j).value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase1":
        for j in range(1, sheet.max_column + 1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict)
